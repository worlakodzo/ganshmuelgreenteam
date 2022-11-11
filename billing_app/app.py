from flask import Flask, jsonify, request, abort
from db import connection
import os.path
from openpyxl import Workbook, load_workbook
import datetime
import requests


app = Flask(__name__)

 
@app.route('/billing-api/health')
def health_db_status():
    
    try:
        
        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)
            stmt = "select 1"
            mycursor.execute(stmt)
            stmt_result = mycursor.fetchone()
            return jsonify({"status": "OK"}), 200

    except Exception as e:
        
        return jsonify({"status":"failure"}), 500
    

@app.route('/', methods = ["GET", "POST", "PUT"])
def billing_index():

    try:

        data = {
            "billing_index" "This is the home page"
        }


        return jsonify(data)

    except Exception as err:
        abort(500)


@app.route('/provider', methods=['GET', 'POST', 'PUT'])
def provider():
    if request.method == 'POST':
        body = request.get_json()
        id = body['id']
        name = body['name']
        if id != '' and name != '':
            with connection.cursor() as provider:
                provider = connection.cursor(dictionary=True)
                do = "INSERT INTO Provider (`name`) VALUES (%s)"
                provider.execute(do)
                connection.commit()
                return jsonify(id), 201
        else:
            return jsonify({"msg": " Unsuccessfull!!!"}), 204
            
    else:
        with connection.cursor() as provider:
            do = "SELECT * FROM Provider"
            provider.execute(do)
            result = provider.fetchall()
            return jsonify(result)


@app.route('/provider/<id>', methods=['GET', 'POST', 'PUT'])
def update_provider_name(id):
    if request.method == 'PUT':
        body = request.args.get()
        name = body['name']
        if name != '':
            with connection.cursor() as provider:
                provider = connection.cursor(dictionary=True)
                do = "INSERT INTO Provider (`name`) VALUES (%s)"
                provider.execute(do)
                connection.commit()
                return jsonify(name), 201
        else:
            return jsonify({"msg": " Unsuccessfull!!!"}), 204            


@app.route('/weight', methods = ["POST"])
def add_weight():



    body = request.get_json()

    data = {
        "container": body['container'],
        "weight": body['weight'],
        "truck_id": body['truck_id'],
        "unit": body['unit'],
        "force": body['force']
    }


    return jsonify(data), 201


@app.route('/rates', methods=['GET', 'POST'])
def rates():
    if request.method == 'POST':
        # receive post parameters
        body = request.get_json()
        # name = body['name']
        file = body['file']
        filepath = f'./in/{file}'
        # check if file exist
        if os.path.isfile(filepath):
            # df = pd.read_excel(file)
            data = []
            df = load_workbook(filepath)
            sheet = df.active
            row_count = sheet.max_row
            for rows in sheet.iter_rows():
                row_cells = []
                for cell in rows:
                    row_cells.append(cell.value)
                data.append(tuple(row_cells))
                
            # establish db connection
            with connection.cursor() as mycursor:
                # mycursor = connection.cursor(dictionary=True)
                for row in data[1:]:
                    mycursor.execute("SELECT * FROM Rates WHERE product_id = %s AND scope = %s", (row[0], row[2], ))
                    result = mycursor.fetchall()
                    if len(result) < 1:
                        mycursor.execute("INSERT INTO Rates (`product_id`, `rate`, `scope`) VALUES (%s, %s, %s)", row)
                    else:
                        mycursor.execute("""
                        UPDATE Rates 
                        SET rate = %s 
                        WHERE product_id = %s AND scope = %s
                        """, (row[1], f'{row[0]}', f'{row[2]}'))
                    connection.commit()
                return jsonify({"msg": "Rates uploaded successfully!"}), 201
        else:
            return jsonify({"msg": "Unsupported file format!!!"}), 204
    else:
        with connection.cursor() as mycursor:
            # mycursor = connection.cursor(dictionary=True)
            stmt = "SELECT * FROM Rates"
            mycursor.execute(stmt)
            stmt_result = mycursor.fetchall()
            # export data to excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "rates"

            ws.append(['Product', 'Rate', 'Scope'])

            for row in stmt_result:
                ws.append(list(row))
            wb.save(f"in/rates-{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
            
            return jsonify({"msg": "Rates downloaded successfully!"})


@app.route('/bill/<id>')
def getbill(id):
    t1 = request.args.get('t1') if request.args.get('t1') else datetime.datetime(2022,1,1).strftime('%Y%m%d%H%M%S')
    t2 = request.args.get('t2') if request.args.get('t2') else datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    param={"from": t1,"to": t2}

    sessionuri = requests.get(f"http://ec2-18-192-110-37.eu-central-1.compute.amazonaws.com:8081/session/{id}").json()
    
    # expected return
    return jsonify({
        "id": sessionuri[0]['id'],
        "name": "<str>",
        "from": "<str>",
        "to": "<str>",
        "truckCount": "<int>",
        "sessionCount": "<int>",
        "products": [{ "product":"<str>","count": "<str>", "amount": "<int>", "rate": "<int>", "pay": "<int>"}],
        "total": sessionuri[0]['neto'] 
    })


#Endpoint for post truck    
@app.route("/truck",methods=['POST'])
def Truck_Post():
        
    if request.method == 'POST':
        body = request.get_json()
        truck_id = body['id']
        provider_id = body['provider_id']
        if truck_id != '' and provider_id != '' :
            
            
             with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                stmt, val = "INSERT INTO Trucks (id, provider_id) VALUES (%s, (select id from Provider where id=%s))", (truck_id, provider_id)
               
               #trap Database Error
                try:            
                    mycursor.execute(stmt, val)
                    connection.commit()
                    return jsonify({"message": "Truck data saved successfully!"}), 201
                except:
                    return jsonify({"msg": "error posting "}), 204
                    
             

        else:
            return jsonify({"msg": "Enter Truck ID and Provider ID "}), 204
    else:
        with connection.cursor() as mycursor:
             mycursor = connection.cursor(dictionary=True)
            # stmt = "SELECT * FROM Truck"
            # mycursor.execute(stmt)
            # stmt_result = mycursor.fetchall()
            # return jsonify(stmt_result)


@app.route("/truck/",methods=['PUT'])
def Truck_Put():
    
    if request.method == 'PUT':
        body = request.get_json()
        truck_id = body['id']
        provider_id = body['provider_id']

        
        if truck_id !='' and provider_id !='':    
            with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                stmt = "update Trucks set provider_id = %s where id=%s" 
                val=(provider_id, truck_id)
                mycursor.execute(stmt,val)
                connection.commit
                return jsonify ({"msg": "provider ID updated successfully! "}+truck_id), 201
    else:
            return jsonify({"msg": "Truck ID not found in the database "}), 204

@app.route('/truck/<id>')
def get_truckid(id):
    t1 = request.args.get('t1')
    t2 = request.args.get('t2')
    param={'id':id,"from":t1,"to":t2}
    reqResp=requests.get("url://weight_server:8081/item/<id>?from=t1&to=t2",params=param)
    assert reqResp.status_code == 200
    data=reqResp.json
    return data


@app.errorhandler(500)
def internal_server_error(error):
    return jsonify({
        "success": False,
        "error": 500,
        "message": "Internal Server Error"
    })


if __name__ == '__main__':
    app.run(debug=True, port=5000)